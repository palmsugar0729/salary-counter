"""openpyxl 导出逻辑，依赖 models 和 config。"""

from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from config import COLOR_PALETTE
from models import MonthData

# A~J 共10列（去掉了原来无意义的 F 空列）
HEADERS = [
    "日期", "星期", "时间", "小时数", "时薪（元/时）",
    "薪酬", "合计时数（小时）", "合计薪酬（元）", "班级", "备注",
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16)
NORMAL_FONT = Font(size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def export(filepath: str | Path, data: MonthData) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工资结算表"

    records = data.sorted_records()
    groups = _build_groups(records)

    # 第1行：标题
    ws.merge_cells("A1:J1")
    ws["A1"] = "工资结算表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # 第2行：姓名
    ws["G2"] = "姓名："
    ws["G2"].font = NORMAL_FONT
    ws["G2"].alignment = LEFT
    ws["H2"] = data.teacher_name
    ws["H2"].font = NORMAL_FONT
    ws["H2"].alignment = LEFT

    # 第3行：表头
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 为每个 group_key 预分配颜色
    color_map = {}
    for idx, gk in enumerate(groups):
        color_map[gk] = COLOR_PALETTE[idx % len(COLOR_PALETTE)]

    # 数据行：按日期+时间顺序输出，同组保持同色（无需连续）
    row_num = 4
    for rec in records:
        color = color_map[rec.group_key]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        _write_record_row(ws, row_num, rec, fill)
        row_num += 1

    # 合计行
    last_data_row = row_num - 1
    if records:
        # G 列(7)：合计时数 = SUM(D4:Dn)
        sum_hours_cell = ws.cell(row=row_num, column=7)
        sum_hours_cell.value = f"=SUM(D4:D{last_data_row})"
        sum_hours_cell.font = Font(bold=True, size=11)
        sum_hours_cell.alignment = CENTER
        sum_hours_cell.border = THIN_BORDER

        # H 列(8)：合计薪酬 = SUM(F4:Fn)
        sum_salary_cell = ws.cell(row=row_num, column=8)
        sum_salary_cell.value = f"=SUM(F4:F{last_data_row})"
        sum_salary_cell.font = Font(bold=True, size=11)
        sum_salary_cell.alignment = CENTER
        sum_salary_cell.border = THIN_BORDER

        for col_idx in range(1, 11):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value is None:
                cell.border = THIN_BORDER

    # 列宽
    widths = [13, 8, 10, 10, 14, 10, 14, 14, 10, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(filepath)


def _write_record_row(ws, row: int, rec, fill: PatternFill) -> None:
    values = [
        rec.date,                              # A: 日期
        rec.weekday_str,                       # B: 星期
        rec.start_time.strftime("%H:%M"),       # C: 时间
        rec.hours,                              # D: 小时数
        rec.rate,                               # E: 时薪
        None,                                   # F: 薪酬（公式）
        None,                                   # G: 合计时数（留空）
        None,                                   # H: 合计薪酬（留空）
        rec.class_type,                         # I: 班级
        rec.note,                               # J: 备注
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx)
        if val is not None:
            cell.value = val
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if col_idx <= 8 else LEFT
        cell.border = THIN_BORDER
        cell.fill = fill

    # F 列(6)：薪酬公式 =D*E
    f_cell = ws.cell(row=row, column=6)
    f_cell.value = f"=D{row}*E{row}"


def _build_groups(records: list) -> list[tuple]:
    """按排序后的首次出现顺序提取 group_key，保持时间顺序。"""
    seen = []
    for rec in records:
        if rec.group_key not in seen:
            seen.append(rec.group_key)
    return seen
